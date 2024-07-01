import json
import os
from fpdf import FPDF
import openpyxl
from openpyxl import Workbook

def save_to_text(employee, filename):
    mode = 'a' if os.path.exists(filename) else 'w'
    with open(filename, mode) as file:
        if mode == 'w':
            file.write("Name, DOB, Phone, Email, Age\n")
        file.write(f"{employee['Name']}, {employee['DOB']}, {employee['Phone']}, {employee['Email']}, {employee['Age']}\n")

def save_to_excel(employee, filename):
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "DOB", "Phone", "Email", "Age"])

    worksheet.append([employee['Name'], employee['DOB'], employee['Phone'], employee['Email'], employee['Age']])
    workbook.save(filename)

def save_to_pdf(employee, filename):
    pdf = FPDF()
    if os.path.exists(filename):
        pdf.add_page()
        pdf.set_font("Arial", size=12)
    else:
        pdf.add_page()
        pdf.set_font("Arial", size=12)
    
    for key, value in employee.items():
        pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
    
    pdf.output(filename)
